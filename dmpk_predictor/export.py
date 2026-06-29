"""
Export prediction results to a cleanly formatted Excel workbook.

These are computed prediction outputs (not a live financial model), so values are
written directly. Formatting follows a professional convention: Arial font,
styled header, frozen header row, grouped/auto-sized columns, sensible number
formats, and red highlighting for rows that errored.
"""
from __future__ import annotations

import io
from datetime import datetime
from typing import Mapping, Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# Preferred column order (present columns are shown in this order; extras appended).
_COLUMN_ORDER = [
    "id", "smiles", "error",
    "mw", "clogp", "logd", "ionisation_class", "tpsa",
    "matrix", "fu_p", "blood_plasma_ratio", "fu_inc",
    "clint_liver_mL_min_kg", "cl_u_int_mL_min_kg",
    "clh_blood_mL_min_kg", "clh_plasma_mL_min_kg",
    "E_H", "Fa", "Fg", "Fh", "F_pct",
    "solubility_uM", "permeability_cm_s", "efflux_ratio",
    "target_type", "target_free", "dose_mg",
    "cl_human_msa", "vd_human_msa", "msa_cl_exponent", "msa_cl_r2",
    "cl_human_rat_sss", "vd_human_rat_sss",
    "cl_human_dog_sss", "vd_human_dog_sss",
    "cl_human_cyno_sss", "vd_human_cyno_sss",
    "cl_human_mouse_sss", "vd_human_mouse_sss",
]

# number format per column
_NUM_FMT = {
    "mw": "0.0", "clogp": "0.00", "logd": "0.00", "tpsa": "0.0",
    "fu_p": "0.0000", "blood_plasma_ratio": "0.00", "fu_inc": "0.0000",
    "clint_liver_mL_min_kg": "0.000", "cl_u_int_mL_min_kg": "0.00",
    "clh_blood_mL_min_kg": "0.000", "clh_plasma_mL_min_kg": "0.000",
    "E_H": "0.000", "Fa": "0.000", "Fg": "0.000", "Fh": "0.000", "F_pct": "0.0",
    "solubility_uM": "0.0", "permeability_cm_s": "0.00E+00", "efflux_ratio": "0.00",
    "target_free": "#,##0.0", "dose_mg": "#,##0.0",
    "cl_human_msa": "0.000", "vd_human_msa": "0.000",
    "msa_cl_exponent": "0.000", "msa_cl_r2": "0.000",
    "cl_human_rat_sss": "0.000", "vd_human_rat_sss": "0.000",
    "cl_human_dog_sss": "0.000", "vd_human_dog_sss": "0.000",
    "cl_human_cyno_sss": "0.000", "vd_human_cyno_sss": "0.000",
    "cl_human_mouse_sss": "0.000", "vd_human_mouse_sss": "0.000",
}

# friendlier header labels
_HEADER = {
    "id": "ID", "smiles": "SMILES", "error": "Error",
    "mw": "MW (g/mol)", "clogp": "cLogP", "logd": "LogD", "tpsa": "TPSA",
    "ionisation_class": "Ion. class", "matrix": "Matrix",
    "fu_p": "fu,p", "blood_plasma_ratio": "B:P", "fu_inc": "fu,inc",
    "clint_liver_mL_min_kg": "CLint,liver (mL/min/kg)",
    "cl_u_int_mL_min_kg": "CLu,int (mL/min/kg)",
    "clh_blood_mL_min_kg": "CLh,blood (mL/min/kg)",
    "clh_plasma_mL_min_kg": "CLh,plasma (mL/min/kg)",
    "E_H": "E_H", "Fa": "Fa", "Fg": "Fg", "Fh": "Fh", "F_pct": "F (%)",
    "solubility_uM": "Solubility (uM)", "permeability_cm_s": "Papp (cm/s)",
    "efflux_ratio": "Efflux ratio",
    "target_type": "Target", "target_free": "Target (free)", "dose_mg": "Dose (mg)",
    "cl_human_msa": "CL MSA (mL/min/kg)", "vd_human_msa": "Vd MSA (L/kg)",
    "msa_cl_exponent": "MSA exponent", "msa_cl_r2": "MSA R²",
    "cl_human_rat_sss": "CL rat-SSS", "vd_human_rat_sss": "Vd rat-SSS",
    "cl_human_dog_sss": "CL dog-SSS", "vd_human_dog_sss": "Vd dog-SSS",
    "cl_human_cyno_sss": "CL cyno-SSS", "vd_human_cyno_sss": "Vd cyno-SSS",
    "cl_human_mouse_sss": "CL mouse-SSS", "vd_human_mouse_sss": "Vd mouse-SSS",
}

_HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
_HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
_BODY_FONT = Font(name="Arial", size=10)
_ERROR_FILL = PatternFill("solid", fgColor="FCE4E4")
_THIN = Side(style="thin", color="D9D9D9")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _ordered_columns(df: pd.DataFrame) -> list[str]:
    present = [c for c in _COLUMN_ORDER if c in df.columns]
    extras = [c for c in df.columns if c not in present]
    return present + extras


def results_to_workbook(df: pd.DataFrame, meta: Optional[Mapping] = None) -> Workbook:
    """Build a formatted openpyxl Workbook from a results DataFrame."""
    cols = _ordered_columns(df)
    wb = Workbook()
    ws = wb.active
    ws.title = "Predictions"

    # header
    for j, col in enumerate(cols, start=1):
        c = ws.cell(row=1, column=j, value=_HEADER.get(col, col))
        c.fill = _HEADER_FILL
        c.font = _HEADER_FONT
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _BORDER

    # body
    for i, (_, row) in enumerate(df.iterrows(), start=2):
        has_error = bool(row.get("error"))
        for j, col in enumerate(cols, start=1):
            val = row[col]
            if pd.isna(val):
                val = None
            c = ws.cell(row=i, column=j, value=val)
            c.font = _BODY_FONT
            c.border = _BORDER
            if col in _NUM_FMT and isinstance(val, (int, float)):
                c.number_format = _NUM_FMT[col]
            if col in ("smiles", "error", "id"):
                c.alignment = Alignment(horizontal="left")
            else:
                c.alignment = Alignment(horizontal="center")
            if has_error:
                c.fill = _ERROR_FILL

    # column widths (auto-ish, clamped)
    for j, col in enumerate(cols, start=1):
        header_len = len(str(_HEADER.get(col, col)))
        try:
            data_len = df[col].astype(str).map(len).max()
        except Exception:
            data_len = 8
        width = min(max(header_len + 2, int(data_len) + 2, 9), 42)
        ws.column_dimensions[get_column_letter(j)].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{len(df) + 1}"

    _add_run_info(wb, df, meta)
    return wb


def _add_run_info(wb: Workbook, df: pd.DataFrame, meta: Optional[Mapping]) -> None:
    ws = wb.create_sheet("Run Info")
    rows = [
        ("Generated", datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("Compounds", len(df)),
        ("Errored", int(df["error"].notna().sum()) if "error" in df else 0),
    ]
    if meta:
        rows += list(meta.items())
    rows += [
        ("", ""),
        ("Method", "Well-stirred IVIVE + mechanistic F (Fa·Fg·Fh); dose from free target."),
        ("Note", "Predictions, not measurements. Validate against the source workbook."),
    ]
    for i, (k, v) in enumerate(rows, start=1):
        a = ws.cell(row=i, column=1, value=k); a.font = Font(name="Arial", bold=True, size=10)
        b = ws.cell(row=i, column=2, value=v); b.font = Font(name="Arial", size=10)
        b.alignment = Alignment(horizontal="left", wrap_text=True)
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 70


def results_to_excel_bytes(df: pd.DataFrame, meta: Optional[Mapping] = None) -> bytes:
    """Return the formatted workbook as bytes (for a Streamlit download button)."""
    wb = results_to_workbook(df, meta)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# --------------------------------------------------------------------------- #
# Worksheet-style report (mirrors the Excel prediction sheet) with a Source col
# --------------------------------------------------------------------------- #
_SECTION_FILL = PatternFill("solid", fgColor="C6E0B4")   # light green, like the sheet
_SRC_FILL = {
    "CDD": PatternFill("solid", fgColor="C6EFCE"),       # green  = pulled from CDD
    "manual": PatternFill("solid", fgColor="DDEBF7"),    # blue   = user override
    "default": PatternFill("solid", fgColor="FFEB9C"),   # amber  = placeholder default
    "NA": PatternFill("solid", fgColor="F8CBAD"),         # red    = not reported in CDD
}
_ARIAL = Font(name="Arial", size=10)
_ARIAL_B = Font(name="Arial", size=10, bold=True)


def _section(ws, row: int, title: str) -> int:
    c = ws.cell(row=row, column=1, value=title)
    c.font = _ARIAL_B
    for col in range(1, 4):
        ws.cell(row=row, column=col).fill = _SECTION_FILL
    return row + 1


def _kv(ws, row: int, label, value, source: Optional[str] = None, num_fmt=None) -> int:
    ws.cell(row=row, column=1, value=label).font = _ARIAL
    vc = ws.cell(row=row, column=2, value=value)
    vc.font = _ARIAL
    if num_fmt and isinstance(value, (int, float)):
        vc.number_format = num_fmt
    if source:
        sc = ws.cell(row=row, column=3, value=source)
        sc.font = _ARIAL
        sc.fill = _SRC_FILL.get(source, _SRC_FILL["default"])
        sc.alignment = Alignment(horizontal="center")
    return row + 1


def worksheet_report_bytes(ctx: Mapping) -> bytes:
    """Build a worksheet-style .xlsx report from the app context.

    ctx keys (all optional): compound_id, smiles, mw, logd, ion_class,
      targets {name: (value, source)},
      invivo {species: {metric: (value, source)}},
      invitro {label: (value, source)},
      selections {label: value},
      predictions {label: value},
      simulation {label: value},
      profile (t_list, free_list, target_value).
    `source` is one of CDD / manual / default / NA.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Worksheet"
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 12
    r = 1

    ws.cell(row=r, column=1, value="DMPK Human Dose Prediction").font = Font(
        name="Arial", size=13, bold=True); r += 1
    ws.cell(row=r, column=1, value=datetime.now().strftime("%Y-%m-%d %H:%M")).font = _ARIAL
    r += 2

    # legend
    r = _section(ws, r, "Source legend")
    for src, desc in (("CDD", "pulled from CDD Vault"), ("manual", "user-entered / override"),
                      ("default", "placeholder default (no data)"),
                      ("NA", "not reported in CDD")):
        sc = ws.cell(row=r, column=1, value=src); sc.fill = _SRC_FILL[src]; sc.font = _ARIAL
        ws.cell(row=r, column=2, value=desc).font = _ARIAL
        r += 1
    r += 1

    r = _section(ws, r, "Compound")
    r = _kv(ws, r, "Compound ID", ctx.get("compound_id", ""))
    r = _kv(ws, r, "SMILES", ctx.get("smiles", ""))
    r = _kv(ws, r, "Molecular weight (g/mol)", ctx.get("mw"), num_fmt="0.0")
    r = _kv(ws, r, "LogD", ctx.get("logd"), num_fmt="0.00")
    r = _kv(ws, r, "Ionisation class", ctx.get("ion_class", ""))
    r += 1

    r = _section(ws, r, "Targets (free conc)")
    for name, (val, src) in ctx.get("targets", {}).items():
        r = _kv(ws, r, name, val, src, num_fmt="#,##0.0")
    r += 1

    r = _section(ws, r, "In vivo PK")
    ws.cell(row=r, column=1, value="Species / metric").font = _ARIAL_B
    ws.cell(row=r, column=2, value="Value").font = _ARIAL_B
    ws.cell(row=r, column=3, value="Source").font = _ARIAL_B
    r += 1
    for sp, metrics in ctx.get("invivo", {}).items():
        for metric, (val, src) in metrics.items():
            r = _kv(ws, r, f"{sp} {metric}", val, src, num_fmt="0.000")
    r += 1

    r = _section(ws, r, "In vitro data")
    for label, (val, src) in ctx.get("invitro", {}).items():
        r = _kv(ws, r, label, val, src, num_fmt="0.000")
    r += 1

    r = _section(ws, r, "Selections")
    for label, val in ctx.get("selections", {}).items():
        r = _kv(ws, r, label, val)
    r += 1

    r = _section(ws, r, "Human predictions")
    for label, val in ctx.get("predictions", {}).items():
        r = _kv(ws, r, label, val, num_fmt="#,##0.000")
    r += 1

    r = _section(ws, r, "Simulation parameters & outputs")
    for label, val in ctx.get("simulation", {}).items():
        r = _kv(ws, r, label, val, num_fmt="0.000")

    assumptions = ctx.get("assumptions")
    if assumptions:
        r += 1
        r = _section(ws, r, "Assumptions & limitations")
        for a in assumptions:
            c = ws.cell(row=r, column=1, value="• " + str(a))
            c.font = _ARIAL
            c.alignment = Alignment(wrap_text=True, vertical="top")
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
            r += 1

    # ---- PK profile sheet + chart ----
    prof = ctx.get("profile")
    if prof:
        t_list, free_list, target = prof
        ps = wb.create_sheet("PK Profile")
        ps.cell(row=1, column=1, value="Time (h)").font = _ARIAL_B
        ps.cell(row=1, column=2, value="Free Cp (nM)").font = _ARIAL_B
        if target:
            ps.cell(row=1, column=3, value="Target (nM)").font = _ARIAL_B
        for i, (t, c) in enumerate(zip(t_list, free_list), start=2):
            ps.cell(row=i, column=1, value=round(float(t), 2))
            ps.cell(row=i, column=2, value=round(float(c), 4))
            if target:
                ps.cell(row=i, column=3, value=target)
        from openpyxl.chart import LineChart, Reference
        chart = LineChart()
        chart.title = "Simulated Human PK"
        chart.x_axis.title = "Time (h)"
        chart.y_axis.title = "Free Plasma Concentration (nM)"
        n = len(t_list) + 1
        data = Reference(ps, min_col=2, max_col=3 if target else 2, min_row=1, max_row=n)
        cats = Reference(ps, min_col=1, min_row=2, max_row=n)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.height, chart.width = 9, 18
        ps.add_chart(chart, "E2")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
