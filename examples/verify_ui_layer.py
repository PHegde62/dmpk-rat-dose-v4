# smiles-ok-file  (contains only public reference SMILES)
"""Verify the formatted-Excel export and the table->ADME row mapping."""
import math
import pandas as pd
from dmpk_predictor.export import results_to_excel_bytes
from dmpk_predictor.pipeline import build_adme_from_row  # noqa: F401

df = pd.DataFrame([
    {"id": "aspirin", "smiles": "CC(=O)Oc1ccccc1C(=O)O", "mw": 180.16, "clogp": 1.31,
     "logd": 1.31, "ionisation_class": "acidic", "tpsa": 63.6, "matrix": "microsome",
     "fu_p": 0.5, "blood_plasma_ratio": 0.9, "fu_inc": 0.8, "clint_liver_mL_min_kg": 22.5,
     "cl_u_int_mL_min_kg": 28.1, "clh_blood_mL_min_kg": 7.9, "clh_plasma_mL_min_kg": 7.1,
     "E_H": 0.34, "Fa": 1.0, "Fg": 1.0, "Fh": 0.66, "F_pct": 66.0, "solubility_uM": 16650.0,
     "permeability_cm_s": 1.2e-5, "target_type": "AUC", "target_free": 5000.0,
     "dose_mg": 1234.5, "error": None},
    {"id": "broken", "smiles": "not_smiles", "error": "SMILES error: Unparseable SMILES"},
])

b = results_to_excel_bytes(df, meta={"Target": "AUC", "tau (h)": 24})
open("examples/_export_test.xlsx", "wb").write(b)

from openpyxl import load_workbook
wb = load_workbook("examples/_export_test.xlsx")
ws = wb["Predictions"]
errs = [c.value for r in ws.iter_rows() for c in r
        if isinstance(c.value, str) and c.value.startswith("#")]
print("xlsx bytes:", len(b), "| sheets:", wb.sheetnames,
      "| freeze:", ws.freeze_panes, "| filter:", ws.auto_filter.ref)
print("excel error cells:", errs)

row = {"smiles": "CCO", "clint": 10, "clint_unit": "uL/min/mg", "matrix": "hepatocyte",
       "fu_p": 50, "fu_p_unit": "% bound", "blood_plasma_ratio": 0.8, "permeability": 120,
       "permeability_unit": "nm/s", "vd_human": 1.5,
       "solubility": float("nan"), "efflux_ratio": float("nan")}
print("adme mapping:", build_adme_from_row(row))
