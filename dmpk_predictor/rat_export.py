"""
Excel export for the V4 batch rat-dose pipeline.

Produces a workbook with three sheets:
  * "Rat Dose Predictions" — one row per compound: descriptors, rat CL / Vss / F,
    the rat maintenance dose (mg and mg/kg), half-life and profile Cmax/Cmin.
  * "PK Profiles" — every compound's free-plasma concentration vs time, side by
    side, with a single overlay LineChart.
  * "Run Info" — generation time, settings, method note and caveats.

Values are computed predictions, so they are written directly (no live formulas),
formatted in the same professional style as the V1-V3 exporter.
"""
from __future__ import annotations

import io
from datetime import datetime
from typing import Mapping, Optional, Sequence

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference, ScatterChart, Series
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# Column order for the predictions sheet (present columns shown in this order).
_COLUMN_ORDER = [
    "id", "smiles", "error", "adme_source",
    "mw", "clogp", "logd", "ionisation_class", "tpsa",
    "cl_source", "matrix", "fu_p", "blood_plasma_ratio", "fu_inc",
    "clint_liver_mL_min_kg", "cl_u_int_mL_min_kg",
    "clh_blood_mL_min_kg", "cl_rat_plasma_mL_min_kg", "cl_rat_obs_mL_min_kg",
    "ivive_pred_over_obs", "ivive_fold_error",
    "E_H", "Fa", "Fg", "Fh", "F_pct",
    "vss_rat_L_kg", "vd_method", "vss_pred_oie_L_kg", "vss_obs_L_kg",
    "vss_pred_over_obs", "vss_fold_error", "kel_per_h", "t_half_h",
    "solubility_uM", "permeability_cm_s",
    "target_type", "target_free_nM", "tau_h",
    "dose_mg", "dose_mg_kg", "profile_cmax_free_nM", "profile_cmin_free_nM",
    "flag",
]

_NUM_FMT = {
    "mw": "0.0", "clogp": "0.00", "logd": "0.00", "tpsa": "0.0",
    "fu_p": "0.0000", "blood_plasma_ratio": "0.00", "fu_inc": "0.0000",
    "clint_liver_mL_min_kg": "0.000", "cl_u_int_mL_min_kg": "0.00",
    "clh_blood_mL_min_kg": "0.000", "cl_rat_plasma_mL_min_kg": "0.000",
    "cl_rat_obs_mL_min_kg": "0.000",
    "ivive_pred_over_obs": "0.00", "ivive_fold_error": "0.00",
    "E_H": "0.000", "Fa": "0.000", "Fg": "0.000", "Fh": "0.000", "F_pct": "0.0",
    "vss_rat_L_kg": "0.000", "vss_pred_oie_L_kg": "0.000", "vss_obs_L_kg": "0.000",
    "vss_pred_over_obs": "0.00", "vss_fold_error": "0.00",
    "kel_per_h": "0.0000", "t_half_h": "0.00",
    "solubility_uM": "0.0", "permeability_cm_s": "0.00E+00",
    "target_free_nM": "#,##0.0", "tau_h": "0.0",
    "dose_mg": "#,##0.0000", "dose_mg_kg": "#,##0.000",
    "profile_cmax_free_nM": "#,##0.00", "profile_cmin_free_nM": "#,##0.00",
}

_HEADER = {
    "id": "ID", "smiles": "SMILES", "error": "Error", "adme_source": "ADME source",
    "mw": "MW (g/mol)", "clogp": "cLogP", "logd": "LogD", "tpsa": "TPSA",
    "ionisation_class": "Ion. class", "cl_source": "CL source", "matrix": "Matrix",
    "fu_p": "rat fu,p", "blood_plasma_ratio": "rat B:P", "fu_inc": "fu,inc",
    "clint_liver_mL_min_kg": "CLint,liver (mL/min/kg)",
    "cl_u_int_mL_min_kg": "CLu,int (mL/min/kg)",
    "clh_blood_mL_min_kg": "CLh,blood (mL/min/kg)",
    "cl_rat_plasma_mL_min_kg": "Rat CL pred (mL/min/kg)",
    "cl_rat_obs_mL_min_kg": "Rat CL obs (mL/min/kg)",
    "ivive_pred_over_obs": "IVIVE pred/obs", "ivive_fold_error": "IVIVE fold-error",
    "E_H": "E_H", "Fa": "Fa", "Fg": "Fg", "Fh": "Fh", "F_pct": "F (%)",
    "vss_rat_L_kg": "Rat Vss (L/kg)", "vd_method": "Vss method",
    "vss_pred_oie_L_kg": "Vss pred ØT (L/kg)", "vss_obs_L_kg": "Vss obs (L/kg)",
    "vss_pred_over_obs": "Vss pred/obs", "vss_fold_error": "Vss fold-error",
    "kel_per_h": "kel (1/h)", "t_half_h": "t½ (h)",
    "solubility_uM": "Solubility (uM)", "permeability_cm_s": "Papp (cm/s)",
    "target_type": "Target", "target_free_nM": "Target free (nM)", "tau_h": "τ (h)",
    "dose_mg": "Rat dose (mg)", "dose_mg_kg": "Rat dose (mg/kg)",
    "profile_cmax_free_nM": "Cmax,free (nM)", "profile_cmin_free_nM": "Cmin,free (nM)",
    "flag": "Flag",
}

_HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
_HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
_BODY_FONT = Font(name="Arial", size=10)
_ERROR_FILL = PatternFill("solid", fgColor="FCE4E4")
_THIN = Side(style="thin", color="D9D9D9")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _ordered_columns(df: pd.DataFrame) -> list[str]:
    present = [c for c in _COLUMN_ORDER if c in df.columns]
    extras = [c for c in df.columns
              if c not in present and not str(c).startswith("_")
              and c not in ("species", "target_free", "profile_error")]
    return present + extras


def _write_predictions(ws, df: pd.DataFrame) -> None:
    cols = _ordered_columns(df)
    for j, col in enumerate(cols, start=1):
        c = ws.cell(row=1, column=j, value=_HEADER.get(col, col))
        c.fill = _HEADER_FILL
        c.font = _HEADER_FONT
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _BORDER
    for i, (_, row) in enumerate(df.iterrows(), start=2):
        has_error = bool(row.get("error")) and pd.notna(row.get("error"))
        for j, col in enumerate(cols, start=1):
            val = row[col]
            if pd.isna(val):
                val = None
            c = ws.cell(row=i, column=j, value=val)
            c.font = _BODY_FONT
            c.border = _BORDER
            if col in _NUM_FMT and isinstance(val, (int, float)):
                c.number_format = _NUM_FMT[col]
            c.alignment = Alignment(horizontal="left" if col in ("smiles", "error", "id", "vd_method") else "center")
            if has_error:
                c.fill = _ERROR_FILL
    for j, col in enumerate(cols, start=1):
        header_len = len(str(_HEADER.get(col, col)))
        try:
            data_len = df[col].astype(str).map(len).max()
        except Exception:
            data_len = 8
        ws.column_dimensions[get_column_letter(j)].width = min(max(header_len + 2, int(data_len) + 2, 9), 44)
    ws.freeze_panes = "C2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{len(df) + 1}"


def _write_profiles(wb: Workbook, profiles: Sequence[Mapping], target_nM: Optional[float]) -> None:
    if not profiles:
        return
    ws = wb.create_sheet("PK Profiles")
    # shared time grid from the first profile (all profiles share dt / t_end)
    t = profiles[0]["t_h"]
    ws.cell(row=1, column=1, value="Time (h)").font = Font(name="Arial", bold=True, size=10)
    for i, tv in enumerate(t, start=2):
        ws.cell(row=i, column=1, value=tv)
    col = 2
    for prof in profiles:
        ws.cell(row=1, column=col, value=str(prof["id"])).font = Font(name="Arial", bold=True, size=10)
        free = prof["free_nM"]
        for i, cv in enumerate(free, start=2):
            ws.cell(row=i, column=col, value=cv)
        col += 1
    target_col = None
    if target_nM:
        target_col = col
        ws.cell(row=1, column=target_col, value="Target (free nM)").font = Font(name="Arial", bold=True, size=10)
        for i in range(2, len(t) + 2):
            ws.cell(row=i, column=target_col, value=target_nM)

    n_rows = len(t) + 1
    chart = LineChart()
    chart.title = "Simulated rat free plasma concentration"
    chart.x_axis.title = "Time (h)"
    chart.y_axis.title = "Free Cp (nM)"
    last_series_col = target_col or (col - 1)
    data = Reference(ws, min_col=2, max_col=last_series_col, min_row=1, max_row=n_rows)
    cats = Reference(ws, min_col=1, min_row=2, max_row=n_rows)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height, chart.width = 11, 22
    ws.add_chart(chart, f"{get_column_letter(last_series_col + 2)}2")
    ws.column_dimensions["A"].width = 10


def _write_ivive(wb: Workbook, df: pd.DataFrame) -> None:
    """IVIVE breakdown sheet + predicted-vs-observed CL and Vss correlation charts."""
    cols = [
        ("id", "ID"), ("cl_source", "CL source"),
        ("clint_liver_mL_min_kg", "CLint,liver (mL/min/kg)"),
        ("cl_u_int_mL_min_kg", "CLu,int (mL/min/kg)"),
        ("clh_blood_mL_min_kg", "CLh,blood (mL/min/kg)"),
        ("cl_rat_plasma_mL_min_kg", "Rat CL pred (mL/min/kg)"),
        ("cl_rat_obs_mL_min_kg", "Rat CL obs (mL/min/kg)"),
        ("ivive_pred_over_obs", "CL pred/obs"),
        ("ivive_fold_error", "CL fold-error"),
        ("E_H", "E_H"),
        ("vss_pred_oie_L_kg", "Vss pred ØT (L/kg)"),
        ("vss_obs_L_kg", "Vss obs (L/kg)"),
        ("vss_pred_over_obs", "Vss pred/obs"),
        ("vss_fold_error", "Vss fold-error"),
    ]
    cols = [(k, h) for k, h in cols if k in df.columns]
    ws = wb.create_sheet("IVIVE")
    for j, (_, h) in enumerate(cols, start=1):
        c = ws.cell(row=1, column=j, value=h)
        c.fill = _HEADER_FILL; c.font = _HEADER_FONT; c.border = _BORDER
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for i, (_, row) in enumerate(df.iterrows(), start=2):
        for j, (k, _) in enumerate(cols, start=1):
            v = row.get(k)
            if pd.isna(v):
                v = None
            c = ws.cell(row=i, column=j, value=v)
            c.font = _BODY_FONT; c.border = _BORDER
            if k in _NUM_FMT and isinstance(v, (int, float)):
                c.number_format = _NUM_FMT[k]
            c.alignment = Alignment(horizontal="left" if k in ("id", "cl_source") else "center")
    for j, (_, h) in enumerate(cols, start=1):
        ws.column_dimensions[get_column_letter(j)].width = min(max(len(h) + 2, 12), 26)

    def _corr(title, obs_col, pred_col, base_col, anchor, unit):
        """Write an obs/pred/unity data block and a predicted-vs-observed scatter."""
        if not {obs_col, pred_col}.issubset(df.columns):
            return
        corr = df.dropna(subset=[obs_col, pred_col])
        if len(corr) < 1:
            return
        start = len(df) + 4
        bc = base_col
        ws.cell(row=start, column=bc, value=title).font = Font(name="Arial", bold=True, size=10)
        ws.cell(row=start + 1, column=bc, value="Obs").font = Font(name="Arial", bold=True, size=9)
        ws.cell(row=start + 1, column=bc + 1, value="Pred").font = Font(name="Arial", bold=True, size=9)
        ws.cell(row=start + 1, column=bc + 2, value="Unity").font = Font(name="Arial", bold=True, size=9)
        obs = corr[obs_col].astype(float).tolist()
        pred = corr[pred_col].astype(float).tolist()
        for r, (o, p) in enumerate(zip(obs, pred), start=start + 2):
            ws.cell(row=r, column=bc, value=round(o, 4))
            ws.cell(row=r, column=bc + 1, value=round(p, 4))
            ws.cell(row=r, column=bc + 2, value=round(o, 4))  # y=x reference
        n0, n1 = start + 2, start + 1 + len(corr)
        chart = ScatterChart()
        chart.title = title
        chart.x_axis.title = f"Observed ({unit})"
        chart.y_axis.title = f"Predicted ({unit})"
        xref = Reference(ws, min_col=bc, min_row=n0, max_row=n1)
        s_pred = Series(Reference(ws, min_col=bc + 1, min_row=n0, max_row=n1), xref, title="pred vs obs")
        s_pred.marker.symbol = "circle"; s_pred.graphicalProperties.line.noFill = True
        s_unity = Series(Reference(ws, min_col=bc + 2, min_row=n0, max_row=n1), xref, title="unity (y=x)")
        chart.series.append(s_pred); chart.series.append(s_unity)
        chart.height, chart.width = 9, 13
        ws.add_chart(chart, anchor)

    n_anchor = len(cols) + 2
    _corr("IVIVE: predicted vs observed rat CL", "cl_rat_obs_mL_min_kg",
          "cl_rat_plasma_mL_min_kg", 1, f"{get_column_letter(n_anchor)}2", "mL/min/kg")
    _corr("Vss: Øie–Tozer predicted vs observed rat Vss", "vss_obs_L_kg",
          "vss_pred_oie_L_kg", 5, f"{get_column_letter(n_anchor)}20", "L/kg")


def _write_run_info(wb: Workbook, df: pd.DataFrame, meta: Optional[Mapping]) -> None:
    ws = wb.create_sheet("Run Info")
    n_err = int(df["error"].notna().sum()) if "error" in df else 0
    rows = [
        ("Generated", datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("Tool", "DMPK Rat Dose Predictor V4 — batch"),
        ("Compounds", len(df)),
        ("Errored", n_err),
    ]
    if meta:
        rows += list(meta.items())
    rows += [
        ("", ""),
        ("Method", "Rat well-stirred IVIVE (Qh=80 mL/min/kg, rat liver scalars) + "
                   "mechanistic F (Fa·Fg·Fh); rat dose from a free target concentration. "
                   "Body weight 0.25 kg."),
        ("Vss", "Measured rat Vss (rat IV PK) when available; else Øie–Tozer with rat "
                "physiological volumes (Vr=0.364 L/kg, Waters & Lombardo DMD 2010)."),
        ("Caveat", "Predictions, not measurements. Hepatic-CL only (no renal/biliary/"
                   "transporter CL). Validate per program."),
    ]
    for i, (k, v) in enumerate(rows, start=1):
        a = ws.cell(row=i, column=1, value=k); a.font = Font(name="Arial", bold=True, size=10)
        b = ws.cell(row=i, column=2, value=v); b.font = Font(name="Arial", size=10)
        b.alignment = Alignment(horizontal="left", wrap_text=True, vertical="top")
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 86


def rat_batch_to_workbook(df: pd.DataFrame, profiles: Sequence[Mapping],
                          meta: Optional[Mapping] = None) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Rat Dose Predictions"
    _write_predictions(ws, df)
    target_nM = None
    if "target_free_nM" in df.columns and len(df):
        try:
            target_nM = float(df["target_free_nM"].dropna().iloc[0])
        except (IndexError, ValueError):
            target_nM = None
    _write_profiles(wb, profiles, target_nM)
    _write_ivive(wb, df)
    _write_run_info(wb, df, meta)
    return wb


def rat_batch_to_excel_bytes(df: pd.DataFrame, profiles: Sequence[Mapping],
                             meta: Optional[Mapping] = None) -> bytes:
    buf = io.BytesIO()
    rat_batch_to_workbook(df, profiles, meta).save(buf)
    return buf.getvalue()


def rat_batch_to_excel(df: pd.DataFrame, profiles: Sequence[Mapping],
                       path: str, meta: Optional[Mapping] = None) -> str:
    rat_batch_to_workbook(df, profiles, meta).save(path)
    return path
